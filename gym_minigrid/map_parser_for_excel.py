"""
Excel map parser for ASIST Saturn maps.

This module parses two maps contained in `data/map_excel/asist_map.xlsx` and
produces two numpy arrays for MiniGrid-compatible maps. The expected sheet
names are `SaturnA_2.3` and `SaturnB_2.3`.

Changes from the legacy parser:
- Switch to `openpyxl` to read modern .xlsx files (xlrd dropped xlsx support).
- Support for new symbols: A/B/C victims, X collapse plate/threat collapse,
  D falling rubble, P victim detection plate, R/RR/RRR rubble layers,
  T freezing threat, F objects.
- Grey and brown filled cells are considered walls.
- No doors are produced anymore.
- Map extents are larger, spanning coordinates from (-2226,-13) in the
  top-left to (-2087, 64) in the bottom-right. This results in a fixed grid
  of height 78 and width 140.

Output files are saved under `gym_minigrid/envs/resources/` with names:
- `SaturnA_2_3.npy`
- `SaturnB_2_3.npy`
- `raw_map_state_saturnA.npy`, `raw_map_state_saturnB.npy` (walkable==1, walls==4)

All functions are annotated with type hints and in-line comments describe
non-obvious implementation details.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import numpy as np
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import column_index_from_string
try:
    # Older/newer openpyxl versions expose COLOR_INDEX here; fallback to empty
    from openpyxl.styles.colors import COLOR_INDEX  # type: ignore
except Exception:  # pragma: no cover - best-effort compatibility
    COLOR_INDEX = {}  # type: ignore


# Directories
PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_ROOT / "data" / "map_excel"
RESOURCES_DIR = (Path(__file__).parent / "./envs/resources").resolve()

# Workbook configuration
WORKBOOK_PATH = DATA_DIR / "asist_map.xlsx"


@dataclass(frozen=True)
class MapConfig:
    """Configuration describing how to assemble a single Saturn map."""

    name: str
    sheet_name: str
    csv_filename: str
    raw_state_basename: str

    @property
    def csv_path(self) -> Path:
        return DATA_DIR / self.csv_filename


@dataclass(frozen=True)
class MapBlock:
    """Row parsed from a MapBlocks CSV file."""

    x: int
    y: int
    z: int
    block_type: str
    feature_type: str


MAP_CONFIGS: Tuple[MapConfig, ...] = (
    MapConfig(
        name="SaturnA_2_3",
        sheet_name="SaturnA_2.3",
        csv_filename="MapBlocks_SaturnA_2.3_xyz.csv",
        raw_state_basename="raw_map_state_saturna",
    ),
    MapConfig(
        name="SaturnB_2_3",
        sheet_name="SaturnB_2.3",
        csv_filename="MapBlocks_SaturnB_2.3_xyz.csv",
        raw_state_basename="raw_map_state_saturnb",
    ),
    # SaturnC and SaturnD reuse SaturnA/B structural layouts respectively; only
    # their interactive features differ, and those are injected from CSV.
    MapConfig(
        name="SaturnC_2_3",
        sheet_name="SaturnA_2.3",
        csv_filename="MapBlocks_SaturnC_2.3_xyz.csv",
        raw_state_basename="raw_map_state_saturnc",
    ),
    MapConfig(
        name="SaturnD_2_3",
        sheet_name="SaturnB_2.3",
        csv_filename="MapBlocks_SaturnD_2.3_xyz.csv",
        raw_state_basename="raw_map_state_saturnd",
    ),
)

# Coordinate system from Excel range C5:EK80 (1-based). We detect the
# content origin, but for exact control we set explicit offsets:
# C=3, row 5 → origin; EK is column 5*26+?; use openpyxl to compute.
EXCEL_RANGE_START = (5, 3)   # (row, col) = (5, 'C')
EXCEL_RANGE_END = (80, None) # end row; end col resolved per sheet
TOP_LEFT: Tuple[int, int] = (-2225, -11)
BOTTOM_RIGHT: Tuple[int, int] = (-2087, 64)

# Derived grid size (height x width)
GRID_HEIGHT: int = BOTTOM_RIGHT[1] - TOP_LEFT[1] + 1
GRID_WIDTH: int = BOTTOM_RIGHT[0] - TOP_LEFT[0] + 1


# MiniGrid integer IDs used in this repository (see gym_minigrid/index_mapping.py)
EMPTY = 1
WALL = 4            # Default wall
WALL_HEAVY = 30     # Heavy wall
WALL_LIGHT = 31     # Light grey rubble wall
LAVA = 9            # Hazard (red)
BOX = 255           # Generic object (dark brown)
BOX_LIGHT_BLUE = 11 # Light-blue plate (P)
BOX_DARK_BLUE = 12  # Dark-blue object (D)
BOX_RED = 13        # Red object if needed (not used for X)
GOAL_A = 81         # Victim A (green)
GOAL_B = 82         # Victim B (light green)
GOAL_C = 83         # Victim C (yellow)


# Colors used in the Excel for wall identification via fill
# These are common RGB hex strings (uppercase) for grey/brown families.
GREY_HEXES: Iterable[str] = {
    "FF808080", "FF7F7F7F", "FFC0C0C0", "FFBFBFBF", "FFB0B0B0",
    "FF999999", "FF666666", "FF4D4D4D", "FFD9D9D9"
}
BROWN_HEXES: Iterable[str] = {
    "FFCBA986", "FFA8947D", "FF8B4513", "FFA0522D", "FFCD853F",
    "FF7B3F00", "FF5C4033", "FF6F4E37"
}

# Any non-white, non-empty fill should be considered as structure unless
# explicitly mapped by a token; use this to catch additional wall colors.
NON_WALL_EXCEPTIONS: Iterable[str] = {
    # pure white and none
    "FFFFFFFF", "FF000000", "00000000"
}

# Dynamic exceptions collected from the sheet for cells that should be empty
# even if they have a non-white fill. We track both ARGB strings and a more
# robust color signature (type, rgb, indexed, theme, tint) to match theme/tinted
# colors that may not share the exact RGB.
ADDITIONAL_EMPTY_ARGB: set[str] = set()
ADDITIONAL_EMPTY_SIGS: set[Tuple[Any, Any, Any, Any, Any]] = set()

# Dynamic additional wall colors (treated as walls across the sheet)
ADDITIONAL_WALL_ARGB: set[str] = set()
ADDITIONAL_WALL_SIGS: set[Tuple[Any, Any, Any, Any, Any]] = set()
ADDITIONAL_WALL_RGBS: List[Tuple[int, int, int]] = []
ADDITIONAL_WALL_THEME_IDS: set[int] = set()
ADDITIONAL_WALL_INDEXED_IDS: set[int] = set()


def _color_signature(cell: Cell) -> Optional[Tuple[Any, Any, Any, Any, Any]]:
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return None
    color = fill.start_color
    if color is None:
        return None
    return (getattr(color, "type", None), getattr(color, "rgb", None), getattr(color, "indexed", None), getattr(color, "theme", None), getattr(color, "tint", None))


def _rgb_triplet(argb: Optional[str]) -> Optional[Tuple[int, int, int]]:
    if not argb or len(argb) not in (6, 8):
        return None
    # Strip alpha if present (ARGB → RGB)
    hexrgb = argb[-6:]
    try:
        r = int(hexrgb[0:2], 16)
        g = int(hexrgb[2:4], 16)
        b = int(hexrgb[4:6], 16)
        return (r, g, b)
    except Exception:
        return None


def _rgb_distance(c1: Tuple[int, int, int], c2: Tuple[int, int, int]) -> int:
    return abs(c1[0]-c2[0]) + abs(c1[1]-c2[1]) + abs(c1[2]-c2[2])


def _color_to_argb(cell: Cell) -> Optional[str]:
    """Return the ARGB hex string for the cell fill color if present.

    openpyxl may encode colors as RGB, indexed palette, or theme-based.
    This helper normalizes to an ARGB string when possible.
    """
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return None

    color = fill.start_color
    if color is None:
        return None

    # RGB specified directly
    if getattr(color, "type", None) == "rgb" and color.rgb:
        return color.rgb.upper()

    # Indexed palette
    if getattr(color, "type", None) == "indexed" and color.indexed is not None:
        try:
            argb = COLOR_INDEX[color.indexed]
            return (argb or "").upper() or None
        except Exception:
            return None

    # Theme-based colors are hard to resolve without workbook theme; ignore
    return None


def _is_wall_fill(cell: Cell) -> bool:
    """Heuristic: consider filled grey/brown cells as walls.

    Returns True if the cell's fill color matches one of the configured
    grey/brown ARGB values.
    """
    argb = _color_to_argb(cell)
    if not argb:
        return False
    # Dynamic WALL has highest priority
    sig = _color_signature(cell)
    if (argb in ADDITIONAL_WALL_ARGB) or (sig and sig in ADDITIONAL_WALL_SIGS):
        return True
    # Dynamic EMPTY has next priority (prevent palette-based wall)
    if (argb in ADDITIONAL_EMPTY_ARGB) or (sig and sig in ADDITIONAL_EMPTY_SIGS):
        return False
    # Static palettes
    if (argb in GREY_HEXES) or (argb in BROWN_HEXES):
        return True
    # Fuzzy RGB distance to capture slight theme/tint differences
    rgb = _rgb_triplet(argb)
    if rgb and any(_rgb_distance(rgb, ref) <= 48 for ref in ADDITIONAL_WALL_RGBS):
        return True
    # Theme/indexed id match
    color = cell.fill.start_color if cell.fill else None
    if color is not None:
        if getattr(color, "type", None) == "theme" and getattr(color, "theme", None) in ADDITIONAL_WALL_THEME_IDS:
            return True
        if getattr(color, "type", None) == "indexed" and getattr(color, "indexed", None) in ADDITIONAL_WALL_INDEXED_IDS:
            return True
    return False


def _normalize_token(val: Optional[str]) -> str:
    """Normalize cell textual content for symbol matching.

    - Strips whitespace
    - Upper-cases ASCII letters
    - Returns empty string for None
    """
    if val is None:
        return ""
    return str(val).strip().upper()


def symbol_to_minigrid(token: str) -> int:
    """Map a normalized token to a MiniGrid integer id.

    The Excel sheets now only contribute structural information (walls/rubble)
    and static props (``F`` blocks). All interactive entities (victims, signals,
    hazards) are sourced from the MapBlocks CSV overlay.
    """
    if token == "F":  # object → light brown
        return BOX
    if token in {"RRR", "RR", "R"}:  # rubble → light grey wall
        return WALL_LIGHT
    return EMPTY


def find_content_origin(sheet) -> Tuple[int, int]:
    """Best-effort detection of the first map row/column.

    Many mapping sheets include header rows/columns. We scan top-left area
    to find the earliest row and column that contain either a known symbol
    or a filled wall cell.

    Returns (row_start, col_start) as 1-based indices for openpyxl.
    """
    max_scan_rows = min(50, sheet.max_row)
    max_scan_cols = min(50, sheet.max_column)

    row_start = None
    col_start = None

    for r in range(1, max_scan_rows + 1):
        for c in range(1, max_scan_cols + 1):
            cell = sheet.cell(row=r, column=c)
            token = _normalize_token(cell.value)
            if token in {"A", "B", "C", "X", "D", "P", "R", "RR", "RRR", "T", "F"} or _is_wall_fill(cell):
                row_start = r if row_start is None else min(row_start, r)
                col_start = c if col_start is None else min(col_start, c)
    if row_start is None or col_start is None:
        # Fallback to 1,1 if detection fails
        row_start = 1
        col_start = 1
    # Override by fixed range origin C5 if present
    return EXCEL_RANGE_START[0], EXCEL_RANGE_START[1]


_BASE_GRID_CACHE: Dict[str, np.ndarray] = {}


def build_base_grid(sheet_name: str) -> np.ndarray:
    """Parse a Saturn Excel sheet into a base MiniGrid numpy array.

    The returned grid only contains structural elements coming from the sheet
    (walls, rubble, object markers). Dynamic gameplay entities are added later
    from the MapBlocks CSV overlay.
    """
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {WORKBOOK_PATH}")
    sheet = wb[sheet_name]

    # Populate dynamic empty-color exceptions from designated cells
    # As requested: use colors from AT11, CB11, DJ11 as canonical EMPTY colors
    from_cells = ("AT", "CB", "DJ")
    ADDITIONAL_EMPTY_ARGB.clear()
    ADDITIONAL_EMPTY_SIGS.clear()
    ADDITIONAL_WALL_ARGB.clear()
    ADDITIONAL_WALL_SIGS.clear()
    ADDITIONAL_WALL_THEME_IDS.clear()
    ADDITIONAL_WALL_INDEXED_IDS.clear()
    # IMPORTANT: reset RGB reference list per sheet to avoid leakage across runs
    ADDITIONAL_WALL_RGBS.clear()
    for col in from_cells:
        try:
            ci = column_index_from_string(col)
            # sample a few nearby rows to capture theme/tint variants
            for rr in (11, 12, 10):
                cell = sheet.cell(row=rr, column=ci)
                argb = _color_to_argb(cell)
                if argb:
                    ADDITIONAL_EMPTY_ARGB.add(argb)
                sig = _color_signature(cell)
                if sig:
                    ADDITIONAL_EMPTY_SIGS.add(sig)
        except Exception:
            pass

    # Additionally: sample AG10 as a canonical WALL color and include it
    try:
        ag_col = column_index_from_string('AG')
        for rr in (9, 10, 11, 12):
            cell = sheet.cell(row=rr, column=ag_col)
            argb = _color_to_argb(cell)
            if argb:
                ADDITIONAL_WALL_ARGB.add(argb)
                rgb = _rgb_triplet(argb)
                if rgb:
                    ADDITIONAL_WALL_RGBS.append(rgb)
            sig = _color_signature(cell)
            if sig:
                ADDITIONAL_WALL_SIGS.add(sig)
            color = cell.fill.start_color if cell.fill else None
            if color is not None:
                if getattr(color, "type", None) == "theme" and getattr(color, "theme", None) is not None:
                    ADDITIONAL_WALL_THEME_IDS.add(getattr(color, "theme", None))
                if getattr(color, "type", None) == "indexed" and getattr(color, "indexed", None) is not None:
                    ADDITIONAL_WALL_INDEXED_IDS.add(getattr(color, "indexed", None))
    except Exception:
        pass

    # Initialize grid (row-major: y, x)
    grid = np.zeros((GRID_HEIGHT, GRID_WIDTH), dtype=np.int32) + EMPTY

    row_start, col_start = find_content_origin(sheet)

    # Compute Excel end column for EK (use openpyxl translator)
    from openpyxl.utils import column_index_from_string
    excel_end_col = column_index_from_string('EK')
    excel_rows = EXCEL_RANGE_END[0] - EXCEL_RANGE_START[0] + 1
    excel_cols = excel_end_col - EXCEL_RANGE_START[1] + 1

    # Validate computed size matches grid size; if not, crop to min
    h = min(GRID_HEIGHT, excel_rows)
    w = min(GRID_WIDTH, excel_cols)

    for r in range(h):
        for c in range(w):
            cell = sheet.cell(row=row_start + r, column=col_start + c)
            token = _normalize_token(cell.value)

            # First, handle dynamic WALL/EMPTY overrides by color
            argb = _color_to_argb(cell)
            sig = _color_signature(cell)
            # Treat any cell matching AG10's color (ARGB/signature/theme/indexed/fuzzy RGB)
            # as WALL prior to any EMPTY override
            if (argb in ADDITIONAL_WALL_ARGB) or (sig and sig in ADDITIONAL_WALL_SIGS):
                grid[r, c] = WALL
                continue
            # Fuzzy RGB/theme/indexed checks for AG10 wall color match
            if argb:
                rgb = _rgb_triplet(argb)
                if (rgb and any(_rgb_distance(rgb, ref) <= 48 for ref in ADDITIONAL_WALL_RGBS)):
                    grid[r, c] = WALL
                    continue
            color = cell.fill.start_color if cell.fill else None
            if color is not None:
                if getattr(color, "type", None) == "theme" and getattr(color, "theme", None) in ADDITIONAL_WALL_THEME_IDS:
                    grid[r, c] = WALL
                    continue
                if getattr(color, "type", None) == "indexed" and getattr(color, "indexed", None) in ADDITIONAL_WALL_INDEXED_IDS:
                    grid[r, c] = WALL
                    continue
            if (argb in ADDITIONAL_EMPTY_ARGB) or (sig and sig in ADDITIONAL_EMPTY_SIGS):
                grid[r, c] = EMPTY
                continue

            # Otherwise: prefer explicit symbol; if still empty, infer by fill
            val = symbol_to_minigrid(token)
            if val == EMPTY and _is_wall_fill(cell):
                val = WALL
            grid[r, c] = val

    # Explicitly enforce the provided world coordinate (-2195, -6) from AG10
    # Map world (x,y) to array indices: col = x - TOP_LEFT[0], row = y - TOP_LEFT[1]
    try:
        enforce_x, enforce_y = -2195, -6
        rr = enforce_y - TOP_LEFT[1]
        cc = enforce_x - TOP_LEFT[0]
        if 0 <= rr < GRID_HEIGHT and 0 <= cc < GRID_WIDTH:
            grid[rr, cc] = WALL
    except Exception:
        pass

    # Add a solid wall border for safety (consistent with legacy behavior)
    grid[0, :] = WALL
    grid[-1, :] = WALL
    grid[:, 0] = WALL
    grid[:, -1] = WALL

    return grid


def _get_base_grid(sheet_name: str) -> np.ndarray:
    """Return a copy of the cached base grid for the given sheet."""
    if sheet_name not in _BASE_GRID_CACHE:
        _BASE_GRID_CACHE[sheet_name] = build_base_grid(sheet_name)
    return _BASE_GRID_CACHE[sheet_name].copy()


def _parse_location_xyz(value: str) -> Tuple[int, int, int]:
    """Decode a `LocationXYZ` string such as `-2215 59 58` into integers."""
    parts = value.replace(",", " ").split()
    if len(parts) < 3:
        raise ValueError(f"Invalid LocationXYZ '{value}'")
    x, y, z = (int(parts[0]), int(parts[1]), int(parts[2]))
    return x, y, z


def _world_to_indices(x: int, z: int) -> Optional[Tuple[int, int]]:
    """Convert world coordinates (x, z) into grid row/column indices."""
    row = z - TOP_LEFT[1]
    col = x - TOP_LEFT[0]
    if 0 <= row < GRID_HEIGHT and 0 <= col < GRID_WIDTH:
        return row, col
    return None


def load_map_blocks(csv_path: Path) -> List[MapBlock]:
    """Load MapBlocks rows from the provided CSV file."""
    if not csv_path.exists():
        raise FileNotFoundError(f"MapBlocks CSV not found: {csv_path}")

    entries: List[MapBlock] = []
    with csv_path.open("r", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            location_raw = row.get("LocationXYZ")
            block_type = (row.get("BlockType") or "").strip()
            feature_type = (row.get("FeatureType") or "").strip()
            if not location_raw or not block_type:
                continue
            try:
                x, y, z = _parse_location_xyz(location_raw)
            except ValueError:
                continue
            entries.append(MapBlock(x=x, y=y, z=z, block_type=block_type, feature_type=feature_type))
    return entries


VICTIM_FEATURE_TO_TILE: Dict[str, int] = {
    "victim a": GOAL_A,
    "victim b": GOAL_B,
    "victim c": GOAL_C,
}


def block_entry_to_tile(entry: MapBlock) -> Optional[int]:
    """Translate a MapBlock entry into the MiniGrid tile id to paint."""
    block = entry.block_type.lower()
    feature = entry.feature_type.lower()

    if block == "gravel":
        return None
    if block == "block_signal_victim":
        return BOX_LIGHT_BLUE
    if block == "block_victim_proximity":
        victim_tile = VICTIM_FEATURE_TO_TILE.get(feature)
        return victim_tile if victim_tile is not None else BOX_LIGHT_BLUE
    if block == "block_rubble_collapse":
        return LAVA
    if block.startswith("block_victim_1"):
        return VICTIM_FEATURE_TO_TILE.get(feature)
    return None


def overlay_map_blocks(grid: np.ndarray, blocks: List[MapBlock]) -> None:
    """Mutate the grid in-place by painting tiles from MapBlocks entries."""
    for entry in blocks:
        tile = block_entry_to_tile(entry)
        if tile is None:
            continue
        indices = _world_to_indices(entry.x, entry.z)
        if indices is None:
            continue
        row, col = indices
        grid[row, col] = tile


def build_raw_map_state(grid: np.ndarray) -> np.ndarray:
    """Return a raw walkable-state array mirroring MiniGrid wall semantics."""
    raw = grid.copy().astype(np.int32)
    raw[raw == BOX] = EMPTY
    raw[raw == BOX_LIGHT_BLUE] = EMPTY
    raw[raw == BOX_DARK_BLUE] = EMPTY
    raw[raw == BOX_RED] = EMPTY
    raw[raw == GOAL_A] = EMPTY
    raw[raw == GOAL_B] = EMPTY
    raw[raw == GOAL_C] = EMPTY
    raw[raw == LAVA] = WALL
    raw[raw == WALL_HEAVY] = WALL
    raw[raw == WALL_LIGHT] = WALL
    return raw


def save_map_outputs(grid: np.ndarray, config: MapConfig) -> None:
    """Persist the numpy arrays for a given configuration."""
    RESOURCES_DIR.mkdir(parents=True, exist_ok=True)
    out_map = RESOURCES_DIR / f"{config.name}.npy"
    np.save(out_map, grid)

    raw_path = RESOURCES_DIR / f"{config.raw_state_basename}.npy"
    np.save(raw_path, build_raw_map_state(grid))

    print(f"Saved map: {out_map}")
    print(f"Saved raw state: {raw_path}")


def build_map_from_config(config: MapConfig) -> None:
    """Assemble, overlay, and persist a single Saturn map."""
    grid = _get_base_grid(config.sheet_name)
    blocks = load_map_blocks(config.csv_path)
    overlay_map_blocks(grid, blocks)
    save_map_outputs(grid, config)


def main() -> None:
    """Entry point: parse both Saturn maps and save outputs.

    This function is idempotent and will overwrite existing files.
    """
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    for config in MAP_CONFIGS:
        print(f"Building {config.name} (sheet={config.sheet_name}, csv={config.csv_filename})")
        build_map_from_config(config)


if __name__ == "__main__":
    main()
